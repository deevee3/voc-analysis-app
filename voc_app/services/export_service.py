"""Export service for insights data in multiple formats."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from voc_app.models import Insight


class ExportService:
    """Handles data export in various formats."""

    @staticmethod
    async def export_insights_csv(
        session: AsyncSession,
        filters: dict[str, Any] | None = None,
    ) -> io.StringIO:
        """Export insights to CSV format.
        
        Args:
            session: Database session
            filters: Optional filter parameters
            
        Returns:
            StringIO buffer containing CSV data
        """
        # Build query with filters (reuse logic from insights API)
        query = select(Insight)
        
        # Apply filters if provided
        if filters:
            # TODO: Apply same filter logic as insights.py list_insights
            pass
        
        result = await session.execute(query)
        insights = result.scalars().all()
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=[
                "id",
                "feedback_id",
                "sentiment_score",
                "sentiment_label",
                "summary",
                "journey_stage",
                "urgency_level",
                "created_at",
            ],
        )
        writer.writeheader()
        
        for insight in insights:
            writer.writerow({
                "id": str(insight.id),
                "feedback_id": str(insight.feedback_id),
                "sentiment_score": insight.sentiment_score,
                "sentiment_label": insight.sentiment_label,
                "summary": insight.summary,
                "journey_stage": insight.journey_stage,
                "urgency_level": insight.urgency_level,
                "created_at": insight.created_at.isoformat() if insight.created_at else None,
            })
        
        output.seek(0)
        return output

    @staticmethod
    async def export_insights_json(
        session: AsyncSession,
        filters: dict[str, Any] | None = None,
    ) -> str:
        """Export insights to JSON format.
        
        Args:
            session: Database session
            filters: Optional filter parameters
            
        Returns:
            JSON string
        """
        query = select(Insight)
        
        if filters:
            # TODO: Apply same filter logic as insights.py list_insights
            pass
        
        result = await session.execute(query)
        insights = result.scalars().all()
        
        data = []
        for insight in insights:
            data.append({
                "id": str(insight.id),
                "feedback_id": str(insight.feedback_id),
                "sentiment_score": float(insight.sentiment_score) if insight.sentiment_score else None,
                "sentiment_label": insight.sentiment_label,
                "summary": insight.summary,
                "pain_points": insight.pain_points,
                "feature_requests": insight.feature_requests,
                "competitor_mentions": insight.competitor_mentions,
                "customer_context": insight.customer_context,
                "journey_stage": insight.journey_stage,
                "urgency_level": insight.urgency_level,
                "created_at": insight.created_at.isoformat() if insight.created_at else None,
            })
        
        return json.dumps(data, indent=2)

    @staticmethod
    async def export_insights_excel(
        session: AsyncSession,
        filters: dict[str, Any] | None = None,
    ) -> io.BytesIO:
        """Export insights to Excel format.
        
        Args:
            session: Database session
            filters: Optional filter parameters
            
        Returns:
            BytesIO buffer containing Excel data
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        query = select(Insight)
        
        if filters:
            # TODO: Apply same filter logic as insights.py list_insights
            pass
        
        result = await session.execute(query)
        insights = result.scalars().all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Insights"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        headers = [
            "ID",
            "Feedback ID",
            "Sentiment Score",
            "Sentiment Label",
            "Summary",
            "Journey Stage",
            "Urgency Level",
            "Created At",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_num, insight in enumerate(insights, 2):
            ws.cell(row=row_num, column=1, value=str(insight.id))
            ws.cell(row=row_num, column=2, value=str(insight.feedback_id))
            ws.cell(row=row_num, column=3, value=float(insight.sentiment_score) if insight.sentiment_score else None)
            ws.cell(row=row_num, column=4, value=insight.sentiment_label)
            ws.cell(row=row_num, column=5, value=insight.summary)
            ws.cell(row=row_num, column=6, value=insight.journey_stage)
            ws.cell(row=row_num, column=7, value=insight.urgency_level)
            ws.cell(row=row_num, column=8, value=insight.created_at.isoformat() if insight.created_at else None)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
